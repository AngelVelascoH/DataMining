import os
from openpyxl import load_workbook
from openpyxl.descriptors import base
from openpyxl.descriptors.base import DateTime
import pyodbc
import datetime
folder = r"PPH"

# Encontramos el ID de la estacion


def getId_estacion(estacion):
    cursor.execute(
        f"select ID from Catalogo_Estaciones where Estacion = '{estacion}';")
    while 1:
        row = cursor.fetchone()
        if not row:
            break

        variable = row[0]
    return variable

# Encontramos el ID del elemento


def getId_elemento(elemento):
    cursor.execute(
        f"SELECT ID FROM Catalogo_elementos where Abreviatura = '{elemento}';")
    while 1:
        row = cursor.fetchone()
        if not row:
            break

        variable = row[0]
    return variable

# Buscamos la cantidad de archivos


def buscar_archivos(ruta):
    archivos_texto = []
    archivos = os.listdir(ruta)
    for archivo in archivos:
        if archivo[-5:] == '.xlsx':
            archivos_texto.append(archivo)
    return archivos_texto


# CREAMOS LA CONEXION SEGURA CON SQL
connection = pyodbc.connect(
    'DRIVER={ODBC Driver 17 for SQL Server};SERVER=ANGEL\SQLEXPRESS;DATABASE=PPH_10-19;Trusted_Connection=yes;')
cursor = connection.cursor()


# recorremos los archivos uno por uno.
for archivo in buscar_archivos(folder):
    print(f"año: {archivo[0:4]}")
    wb = load_workbook(folder+"/"+archivo, data_only=True)
    ws = wb[archivo[0:7]]
    año = int(archivo[0:4])
    IDelemento = getId_elemento(str(archivo[4:6]))
    for x in range(53):
        if (año < 2018):
            mes = str(ws['A'+str(x+2)].value)[5:7]
            dia = str(ws['A'+str(x+2)].value)[8:10]
        else:
            mes = str(ws['A'+str(x+2)].value)[3:5]
            dia = str(ws['A'+str(x+2)].value)[0:2]

        if(dia == '' or dia == 'No' or dia == ' '):
            dia = 11
            mes = 11
        semana_registro = datetime.date(
            año, int(mes), int(dia)).isocalendar()[1]
        for letra in 'BCDEFGHIJKLMNOPQ':
            IDestacion = getId_estacion(str(ws[letra+'1'].value))
            medicion = str(ws[letra+str(x+2)].value)
            if(medicion == '-99' or medicion == '' or medicion == 'None' or medicion == ' '):
                pass
            else:
                tupla = (int(IDelemento), int(año), int(mes), int(dia),
                         int(semana_registro), int(IDestacion), float(medicion))
                with open('PPH_10-19', 'a') as f:
                    f.write(str(tupla)[1:-1]+"\n")

# una vez que se creo el archivo, siguie la carga de datos dentro de la tabla principal.

print("---------------------------------------------")
f = open("PPH_10-19", "r")
print("Insertando los valores en la base de datos...")
contador = 0
with connection.cursor() as cursor2:
    for linea in f:
        print(linea)
        cursor2.execute(f"""INSERT INTO Registros_Principal
                        VALUES({linea});""")
        contador = contador + 1
f.close()
print(f"Todos los registros fueron insertados {contador} lineas afectadas")
# CERRAMOS LA CONEXIÓN CON SQL

cursor.close()
connection.close()
